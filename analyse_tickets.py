import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import datetime

# Read and prepare data
def prepare_data(data):
    # Utilisation de l'encodage cp1252 (Windows) pour les caractères accentués
    df = pd.read_csv(data, encoding='cp1252')
    df['Date de début'] = pd.to_datetime(df['Date de début'])
    df['Date de fermeture'] = pd.to_datetime(df['Date de fermeture'])
    df['Dernière mise à jour'] = pd.to_datetime(df['Dernière mise à jour'])
    
    # Ajout du jour et de la semaine
    df['Jour'] = df['Date de début'].dt.strftime('%Y-%m-%d')
    df['Semaine'] = df['Date de début'].dt.strftime('%Y-S%V')
    df['NumSemaine'] = df['Date de début'].dt.isocalendar().week
    return df

# Create workbook
wb = Workbook()

# 1. Tickets par état
def create_status_sheet(wb, df):
    ws = wb.create_sheet("État des tickets")
    status_count = df['Etat agrégé'].value_counts().reset_index()
    status_count.columns = ['État', 'Nombre de tickets']
    
    for r in dataframe_to_rows(status_count, index=False, header=True):
        ws.append(r)

    chart = PieChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=len(status_count)+1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(status_count)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Distribution des tickets par état"
    
    # Ajout des labels de données
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showPercent = True
    
    ws.add_chart(chart, "E2")

# 2. Tickets par type
def create_category_sheet(wb, df):
    ws = wb.create_sheet("Types de tickets")
    cat_count = df['Sous-classe de Ticket'].value_counts().reset_index()
    cat_count.columns = ['Type', 'Nombre de tickets']
    
    for r in dataframe_to_rows(cat_count, index=False, header=True):
        ws.append(r)

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=len(cat_count)+1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(cat_count)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Tickets par type"
    
    # Ajout des labels de données
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    
    ws.add_chart(chart, "E2")

# 3. Analyse temporelle journalière
def create_time_analysis_sheet(wb, df):
    ws = wb.create_sheet("Analyse journalière")
    
    # Analyse par jour
    daily_data = df.groupby('Jour')['Référence'].count().reset_index()
    daily_data.columns = ['Jour', 'Nombre de tickets']
    
    # Ajout des statistiques par état
    daily_status = df.pivot_table(
        index='Jour',
        columns='Etat agrégé',
        values='Référence',
        aggfunc='count',
        fill_value=0
    ).reset_index()
    
    # Fusion des données
    daily_stats = daily_data.merge(daily_status, on='Jour', how='left')
    
    for r in dataframe_to_rows(daily_stats, index=False, header=True):
        ws.append(r)

    # Graphique linéaire pour l'évolution journalière
    chart = LineChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=len(daily_stats)+1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(daily_stats)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Évolution journalière des tickets"
    
    # Ajout des labels de données
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    
    ws.add_chart(chart, "J2")

    # Graphique en barres pour la répartition par état
    chart2 = BarChart()
    data = Reference(ws, min_col=3, max_col=ws.max_column, min_row=1, max_row=len(daily_stats)+1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(daily_stats)+1)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(labels)
    chart2.title = "États des tickets par jour"
    
    # Ajout des labels de données
    chart2.dataLabels = DataLabelList()
    chart2.dataLabels.showVal = True
    
    ws.add_chart(chart2, "J18")

# 4. Performance des agents
def create_agent_performance_sheet(wb, df):
    ws = wb.create_sheet("Performance agents")
    
    # Stats globales par agent
    agent_stats = df.groupby('Agent->Nom complet').agg({
        'Référence': 'count',
        'Etat agrégé': lambda x: (x == 'Résolu').sum()
    }).reset_index()
    agent_stats.columns = ['Agent', 'Tickets traités', 'Tickets résolus']
    agent_stats['Taux de résolution'] = (agent_stats['Tickets résolus'] / agent_stats['Tickets traités'] * 100).round(2)
    
    for r in dataframe_to_rows(agent_stats, index=False, header=True):
        ws.append(r)

    # Graphique en barres pour les tickets traités et résolus
    chart = BarChart()
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=len(agent_stats)+1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(agent_stats)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Performance des agents"
    
    # Ajout des labels de données
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    
    ws.add_chart(chart, "F2")

# 5. Analyse par client
def create_client_analysis_sheet(wb, df):
    ws = wb.create_sheet("Analyse clients")
    
    # Stats par client
    client_stats = df.groupby('Client->Nom organisation').agg({
        'Référence': 'count',
        'Etat agrégé': lambda x: (x == 'Résolu').sum()
    }).reset_index()
    client_stats.columns = ['Client', 'Total tickets', 'Tickets résolus']
    client_stats['Taux de résolution'] = (client_stats['Tickets résolus'] / client_stats['Total tickets'] * 100).round(2)
    
    for r in dataframe_to_rows(client_stats, index=False, header=True):
        ws.append(r)

    chart = BarChart()
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=len(client_stats)+1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(client_stats)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Tickets par client"
    
    # Ajout des labels de données
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    
    ws.add_chart(chart, "F2")

# 6. Analyse par équipe
def create_team_analysis_sheet(wb, df):
    ws = wb.create_sheet("Analyse équipes")
    
    # Stats par équipe
    team_stats = df.groupby('Equipe->Nom').agg({
        'Référence': 'count',
        'Etat agrégé': lambda x: (x == 'Résolu').sum()
    }).reset_index()
    team_stats.columns = ['Équipe', 'Total tickets', 'Tickets résolus']
    team_stats['Taux de résolution'] = (team_stats['Tickets résolus'] / team_stats['Total tickets'] * 100).round(2)
    
    for r in dataframe_to_rows(team_stats, index=False, header=True):
        ws.append(r)

    chart = PieChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=len(team_stats)+1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(team_stats)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Distribution des tickets par équipe"
    
    # Ajout des labels de données
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showPercent = True
    
    ws.add_chart(chart, "F2")

# 7. Rapport détaillé journalier
def create_detailed_daily_report(wb, df):
    ws = wb.create_sheet("Rapport journalier détaillé")
    
    # En-tête avec style
    header = ['Date', 'Nouveaux tickets', 'Tickets résolus', 'En cours', 'Taux de résolution (%)', 
             'Types principaux', 'Agents les plus actifs', 'Clients principaux']
    ws.append(header)
    
    # Style de l'en-tête
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Analyse par jour
    for jour in sorted(df['Jour'].unique()):
        day_data = df[df['Jour'] == jour]
        
        # Calcul des statistiques
        nouveaux = len(day_data)
        resolus = len(day_data[day_data['Etat agrégé'] == 'Résolu'])
        en_cours = nouveaux - resolus
        taux = round((resolus / nouveaux * 100 if nouveaux > 0 else 0), 2)
        
        # Top types
        top_types = day_data['Sous-classe de Ticket'].value_counts().nlargest(2).index.tolist()
        top_types = ', '.join(top_types)
        
        # Top agents
        top_agents = day_data['Agent->Nom complet'].value_counts().nlargest(2).index.tolist()
        top_agents = ', '.join(top_agents)
        
        # Top clients
        top_clients = day_data['Client->Nom organisation'].value_counts().nlargest(2).index.tolist()
        top_clients = ', '.join(top_clients)
        
        # Ajout de la ligne
        ws.append([
            jour, nouveaux, resolus, en_cours, taux,
            top_types, top_agents, top_clients
        ])
    
    # Ajustement des colonnes
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

# 8. Liste détaillée des tickets
def create_tickets_list(wb, df):
    ws = wb.create_sheet("Liste des tickets")
    
    # Sélection et réorganisation des colonnes
    tickets_df = df[[
        'Référence', 'Date de début', 'Agent->Nom complet', 'Etat agrégé',
        'Sous-classe de Ticket', 'Titre', 'Client->Nom organisation'
    ]].copy()
    
    # Renommage des colonnes pour plus de clarté
    tickets_df.columns = [
        'Référence', 'Date', 'Agent', 'État',
        'Type', 'Description', 'Client'
    ]
    
    # Tri par date décroissante
    tickets_df = tickets_df.sort_values('Date', ascending=False)
    
    # Conversion des dates au format lisible
    tickets_df['Date'] = tickets_df['Date'].dt.strftime('%Y-%m-%d %H:%M')
    
    # Écriture des données
    ws.append(list(tickets_df.columns))
    for r in dataframe_to_rows(tickets_df, index=False, header=False):
        ws.append(r)
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Application des styles à l'en-tête
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Application des styles aux cellules
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if row[3].value == 'Résolu':  # Colonne État
                cell.font = Font(color="006100")  # Vert pour les tickets résolus
            elif row[3].value == 'En cours':
                cell.font = Font(color="C65911")  # Orange pour les tickets en cours
    
    # Ajustement des colonnes
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Figer la première ligne
    ws.freeze_panes = "A2"

# Create summary sheet
def create_summary_sheet(wb, df):
    ws = wb.active
    ws.title = "Résumé"
    
    # Style pour les titres
    title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True)
    
    # Statistiques globales
    summary_data = [
        ["Statistiques globales", ""],
        ["Nombre total de tickets", len(df)],
        ["Tickets résolus", len(df[df['Etat agrégé'] == 'Résolu'])],
        ["Taux de résolution", f"{(len(df[df['Etat agrégé'] == 'Résolu']) / len(df) * 100):.2f}%"],
        ["", ""],
        ["Top performances", ""],
        ["Type de ticket le plus fréquent", df['Sous-classe de Ticket'].mode()[0]],
        ["Agent le plus actif", df['Agent->Nom complet'].mode()[0]],
        ["Client avec le plus de tickets", df['Client->Nom organisation'].mode()[0]],
        ["", ""],
        ["Statistiques du jour", ""],
        ["Nouveaux tickets aujourd'hui", len(df[df['Jour'] == df['Jour'].max()])],
        ["Tickets résolus aujourd'hui", len(df[(df['Jour'] == df['Jour'].max()) & (df['Etat agrégé'] == 'Résolu')])],
    ]
    
    for row in summary_data:
        ws.append(row)
        if row[0] and not row[1]:  # C'est un titre
            for cell in ws[ws.max_row]:
                cell.fill = title_fill
                cell.font = title_font
    
    # Ajustement des colonnes
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

# Main execution
def create_analysis_workbook(data_file, output_file):
    df = prepare_data(data_file)
    
    create_summary_sheet(wb, df)
    create_status_sheet(wb, df)
    create_category_sheet(wb, df)
    create_time_analysis_sheet(wb, df)
    create_agent_performance_sheet(wb, df)
    create_client_analysis_sheet(wb, df)
    create_team_analysis_sheet(wb, df)
    create_detailed_daily_report(wb, df)
    create_tickets_list(wb, df)  # Ajout de la nouvelle feuille
    
    wb.save(output_file)
